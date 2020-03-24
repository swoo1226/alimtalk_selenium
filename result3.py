from bs4 import BeautifulSoup
import requests
import json
import urllib.parse
from seleniumrequests import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import time
import sys
import os
from selenium.webdriver.common.alert import Alert
from openpyxl import Workbook
wb = Workbook()

# absolute_file_path = os.path.abspath("upload.xlsx")
options = Options()
options.headless = True
# browser = webdriver.Chrome(
# executable_path="./chromedriver", options=options)
# browser.get('https://test.pocketsurvey.co.kr/login')


def parse_content(html):
    soup = BeautifulSoup(html, 'html.parser')
    return soup


def open_popup(element):
    # double click을 위한
    actionChains = ActionChains(driver)
    actionChains.move_to_element(element)
    element.click()

    actionChains.double_click(element).perform()
    # ((JavascriptExecutor) driver).executeScript("document.getElementById('map_container').dispatchEvent(new Event('dblclick'));")
    # actionChains.double_click(element).perform()


# 비승인 상태 템플릿 여부 판단 값
hasNotConfirmed = False


# print(sys.argv)
# 1.엠앤와이즈 로그인 페이지 이동
path = "./chromedriver"
driver = webdriver.Chrome(path)
driver.get('https://alimtalk.carrym.com/Gate/')
time.sleep(1)
channel_id_input = driver.find_element_by_id("textfield-1013-inputEl")
# print(channel_id_input)
channel_id_input.send_keys('pocketsurvey')
id_input = driver.find_element_by_id("textfield-1014-inputEl")
id_input.send_keys('admin')
pw_input = driver.find_element_by_id("textfield-1016-inputEl")
pw_input.send_keys("At@pizza1!")
login_btn = driver.find_element_by_id("button-1021")


# 2.마이페이지 이동
login_btn.click()
time.sleep(2)

# 3.알림톡 템플릿 관리 페이지 이동
manage_template = driver.find_element_by_id("ext-element-37")
manage_template.click()
time.sleep(2)

# web_url에 원하는 웹의 URL을 넣어주시면 된다.
# web_url = 'https://alimtalk.carrym.com/Gate/'

# 3-1. 조회 일자 설정 request api
# requestUrl = 'https://alimtalk.carrym.com/getAtTplList.svc'
# request_cookies_browser = driver.get_cookies()
# # s = requests.Session()
# # c = [s.cookies.set(c['name'], c['value']) for c in request_cookies_browser]
# r = requests.post(requestUrl,
#                   data=json.dumps({"UGRP_NM": "", "USER_NM": "", "TPL_CD": "", "TPL_NM": "", "REQ_DTS_ST": sys.argv[1], "REQ_DTS_END": sys.argv[2], "APRV_FG": "", "SENDER_KEY": "165821b5cb4350b2644d90ff2328ee8e13bfd932"}))
# print(r.status_code)
# print(r.text)
# resp = s.post(requestUrl)
# method = post
# Content_type = application/x-www-form-urlencoded; charset=UTF-8
# requestBody = {"UGRP_NM":"","USER_NM":"","TPL_CD":"","TPL_NM":"","REQ_DTS_ST":"20170320","REQ_DTS_END":"20200320","APRV_FG":"","SENDER_KEY":"165821b5cb4350b2644d90ff2328ee8e13bfd932"}

# 4. 알림톡 템플릿 검수 현황 parsing
# 날짜 선택은 datapicker 요소를 직접 눌러야 변경 가능할 듯 하다...
fromDate = driver.find_element_by_id("datefield-1066-inputEl")
fromDatePicker = driver.find_element_by_id("datefield-1066-trigger-picker")
fromDatePickerHeader = driver.find_elements_by_class_name(
    "x-datepicker-header")

driver.execute_script(f"arguments[0].value = '{sys.argv[1]}'", fromDate)
if len(sys.argv) > 2:
    print('has end date')
    toDate = driver.find_element_by_id("datefield-1068-inputEl")
    driver.execute_script(f"arguments[0].value = '{sys.argv[2]}'", toDate)

search = driver.find_element_by_id("button-1084")
search.click()
time.sleep(1)


def getTableElements():
    return driver.find_elements_by_tag_name('table')[1:]


html = driver.page_source
html_doc = parse_content(html)
tables = html_doc.find_all("table")[1:]
tableMatrix_unauthorized = []
tableMatrix_authorized = []
# print(tables)
for tableIdx, table in enumerate(tables):
    # Here you can do whatever you want with the data! You can findAll table row headers, etc...
    # print('table---------', table)
    list_of_rows_unauthorized = []
    list_of_rows_authorized = []
    for rowIdx, row in enumerate(table.findAll('tr')):
        tableElements = getTableElements()
        list_of_cells_unauthorized = []
        list_of_cells_authorized = []
        # print(row)
        # print('--------------------------')
        currRow = row.findAll('td')
        # print('currRow', currRow[1])
        status = currRow[1].findAll('div')
        if(status):
            status = status[0]
        if(len(currRow)):
            if(status.text != '승인'):
                print('검수 미완료 템플릿 발견! 엑셀에 기록할 필요!')
                # 검수 상태는 tr에서 찾은 td들 중 1번째 인덱스 내의 div의 값

                hasNotConfirmed = True
                for cell in currRow:
                    text = cell.findAll('div')[0].text
                    list_of_cells_unauthorized.append(text)
                    # 승인 안 된 템플릿 내용을 보기 위하여 팝업 띄우기까지 완료 상태
                    # 반려된 경우에 어떻게 나타나는지 html 구조가 필요함
                    # print(text)
                # print('tableIdx?', tableIdx)
                # print('list_of_celss', list_of_cells)
                currRow[1].findAll('')
                open_popup(tableElements[tableIdx])
                # currTable = tableElements[tableIdx]
                # print('currTable--------------', currTable)
                # time.sleep(10)

                time.sleep(1)
                # popupSource = driver.page_source
                # popupSource_doc = parse_content(popupSource)
                # xLayers = popupSource_doc.find_all('div', {'class': 'x-layer'})
                # print(xLayers)
                xLayers = driver.find_elements_by_class_name('x-layer')
                # popupTable = xLayers[len(xLayers) - 1].find_all('table')[-2]
                popupTables = xLayers[len(
                    xLayers) - 1].find_elements_by_tag_name('table')[-2]
                # print('popupTable?????', popupTable)
                popupTable = popupTables.find_elements_by_tag_name(
                    'tr')[0].find_elements_by_tag_name('td')
                # print('popupTable???????', popupTable[-1].text)
                list_of_cells_unauthorized.append(popupTable[-1].text)
                close = driver.find_elements_by_class_name(
                    'x-tool-close')[0]
                close.click()
                list_of_rows_unauthorized.append(list_of_cells_unauthorized)
            else:
                print('승인된 템플릿들')
                for cell in currRow:
                    text = cell.findAll('div')[0].text
                    list_of_cells_authorized.append(text)
                currRow[1].findAll('')
                # open_popup(tableElements[tableIdx])
                # time.sleep(1)
                # pageSource = driver.page_source
                # pageSource_doc = parse_content(pageSource)
                # xFieldsetBodyInputs = pageSource_doc.find_all(
                #     'div', {'class': 'x-fieldset-body'})[0].find_all('input')
                # print('xFieldsetBodyInputs!!!!!!!!!!', xFieldsetBodyInputs)
                # # inputValues = xFieldsetBody .find_all('input')
                # list_of_cells_authorized.append()
                # close = driver.find_elements_by_class_name(
                #     'x-tool-close')[0]
                # close.click()
                list_of_rows_authorized.append(list_of_cells_authorized)

    if(list_of_rows_unauthorized):
        tableMatrix_unauthorized.append(list_of_rows_unauthorized)
    if(list_of_rows_authorized):
        tableMatrix_authorized.append(list_of_rows_authorized)

# print(tableMatrix)

ws = wb.active
ws.title = '승인'
ws['A1'] = '순번'
ws['B1'] = '검수상태'
ws['C1'] = '템플릿코드'
ws['D1'] = '템플릿명'
ws['E1'] = '메세지내용'
ws['F1'] = '그룹'
ws['G1'] = '등록자'
ws['H1'] = '등록일'
ws['I1'] = '최종변경일'
ws['J1'] = '비고'
authorized_data = []
for rowIndex, rows in enumerate(tableMatrix_authorized):
    for row in rows:
        print('row???', row)
        ws.append(row)
        authorized_data.append({
            "order": row[0],
            "template_code": row[2],
            "template_name": row[3],
            "template": row[4],
        })
print(authorized_data)
with open("authorized.json", "w", encoding='UTF-8') as json_file:
    json_file.write(json.dumps(authorized_data, ensure_ascii=False))
if(hasNotConfirmed):
    print('승인 안 된 템플릿 존재!')
    ws2 = wb.create_sheet()
    ws2.title = '미승인'
    ws2['A1'] = '순번'
    ws2['B1'] = '검수상태'
    ws2['C1'] = '템플릿코드'
    ws2['D1'] = '템플릿명'
    ws2['E1'] = '메세지내용'
    ws2['F1'] = '그룹'
    ws2['G1'] = '등록자'
    ws2['H1'] = '등록일'
    ws2['I1'] = '최종변경일'
    ws2['J1'] = '비고'
    for rowIndex, rows in enumerate(tableMatrix_unauthorized):
        for row in rows:
            ws2.append(row)
wb.save('result.xlsx')
