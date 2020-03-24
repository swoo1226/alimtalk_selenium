from bs4 import BeautifulSoup
import requests
import json
import urllib.parse
from seleniumrequests import Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import time
import sys
import os
from selenium.webdriver.common.alert import Alert
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# absolute_file_path = os.path.abspath("upload.xlsx")
options = Options()
options.headless = True
# browser = webdriver.Chrome(
# executable_path="./chromedriver", options=options)
# browser.get('https://test.pocketsurvey.co.kr/login')


def parse_content(html):
    soup = BeautifulSoup(html, 'html.parser')
    return soup


# 비승인 상태 템플릿 여부 판단 값
hasNotConfirmed = False


# print(sys.argv)
# 1.엠앤와이즈 로그인 페이지 이동
path = "./chromedriver"
driver = webdriver.Chrome(path)
driver.get('https://alimtalk.carrym.com/Gate/')
time.sleep(1)
channel_id_input = driver.find_element_by_id("textfield-1013-inputEl")
# print(channel_id_input)
channel_id_input.send_keys('pocketsurvey')
id_input = driver.find_element_by_id("textfield-1014-inputEl")
id_input.send_keys('admin')
pw_input = driver.find_element_by_id("textfield-1016-inputEl")
pw_input.send_keys("At@pizza1!")
login_btn = driver.find_element_by_id("button-1021")

# double click을 위한
actionChains = ActionChains(driver)
# 2.마이페이지 이동
login_btn.click()
time.sleep(2)

# 3.알림톡 템플릿 관리 페이지 이동
manage_template = driver.find_element_by_id("ext-element-37")
manage_template.click()
time.sleep(2)

# web_url에 원하는 웹의 URL을 넣어주시면 된다.
# web_url = 'https://alimtalk.carrym.com/Gate/'

# 3-1. 조회 일자 설정 request api
# requestUrl = 'https://alimtalk.carrym.com/getAtTplList.svc'
# request_cookies_browser = driver.get_cookies()
# # s = requests.Session()
# # c = [s.cookies.set(c['name'], c['value']) for c in request_cookies_browser]
# r = requests.post(requestUrl,
#                   data=json.dumps({"UGRP_NM": "", "USER_NM": "", "TPL_CD": "", "TPL_NM": "", "REQ_DTS_ST": sys.argv[1], "REQ_DTS_END": sys.argv[2], "APRV_FG": "", "SENDER_KEY": "165821b5cb4350b2644d90ff2328ee8e13bfd932"}))
# print(r.status_code)
# print(r.text)
# resp = s.post(requestUrl)
# method = post
# Content_type = application/x-www-form-urlencoded; charset=UTF-8
# requestBody = {"UGRP_NM":"","USER_NM":"","TPL_CD":"","TPL_NM":"","REQ_DTS_ST":"20170320","REQ_DTS_END":"20200320","APRV_FG":"","SENDER_KEY":"165821b5cb4350b2644d90ff2328ee8e13bfd932"}
# 4. 알림톡 템플릿 검수 현황 parsing
fromDate = driver.find_element_by_id("datefield-1066-inputEl")
fromDate.click()
driver.execute_script(f"arguments[0].value = '{sys.argv[1]}'", fromDate)
if len(sys.argv) > 2:
    toDate = driver.find_element_by_id("datefield-1068-inputEl")
    driver.execute_script(f"arguments[0].value = '{sys.argv[2]}'", toDate)
search = driver.find_element_by_id("button-1084")
search.click()
time.sleep(1)

tableElements = driver.find_elements_by_tag_name('table')[1:]
html = driver.page_source
html_doc = parse_content(html)
tables = html_doc.find_all("table")[1:]
tableMatrix = []
# print(tables)
for table in tables:
    # Here you can do whatever you want with the data! You can findAll table row headers, etc...
    list_of_rows = []
    for rowIdx, row in enumerate(table.findAll('tr')):
        list_of_cells = []
        # print(row)
        # print('--------------------------')
        currRow = row.findAll('td')
        # print('currRow', currRow[1])
        if(len(currRow)):
            if(currRow[1].findAll('div')[0].text == '반려' or currRow[1].findAll('div')[0].text == '등록'):
                print('검수 미완료 템플릿 발견! 엑셀에 기록할 필요!')
                # 검수 상태는 tr에서 찾은 td들 중 1번째 인덱스 내의 div의 값
                hasNotConfirmed = True
                for cell in currRow:
                    text = cell.findAll('div')[0].text
                    list_of_cells.append(text)
                    print('rowInx?', rowIdx)
                    currTable = tableElements[rowIdx]
                    actionChains.double_click(currTable).perform()
                    time.sleep(1)
                    popupSource = driver.page_source
                    popupSource_doc = parse_content(popupSource)
                    xLayers = popupSource_doc.find_all(
                        'div', {'class': 'x-layer'})
                    # print(xLayers)
                    popupTable = xLayers[len(
                        xLayers) - 1].find_all('table')[-2]
                    # print(popupTable)
                    for table in popupTable:
                        popup_list_of_rows = []
                        for popupRowIdx, popRow in enumerate(table.find_all('tr')):
                            popup_list_of_cells = []
                            popupCurrRow = popRow.find_all('td')
                            print(popupCurrRow)
                            # if(popupCurrRow):
                            #     # print(popupCurrRow)
                    close = driver.find_elements_by_class_name(
                        'x-tool-close')[0]
                    close.click()
                    # 승인 안 된 템플릿 내용을 보기 위하여 팝업 띄우기까지 완료 상태
                    # 반려된 경우에 어떻게 나타나는지 html 구조가 필요함
                    # print(text)
                list_of_rows.append(list_of_cells)
    if(list_of_rows):
        tableMatrix.append(list_of_rows)

print(tableMatrix)
if(hasNotConfirmed):
    print('승인 안 된 템플릿 존재!')
    ws['A1'] = '순번'
    ws['B1'] = '검수상태'
    ws['C1'] = '템플릿코드'
    ws['D1'] = '템플릿명'
    ws['E1'] = '메세지내용'
    ws['F1'] = '그룹'
    ws['G1'] = '등록자'
    ws['H1'] = '등록일'
    ws['I1'] = '최종변경일'
    ws['J1'] = '비고'
    for rowIndex, rows in enumerate(tableMatrix):
        for row in rows:
            ws.append(row)
    wb.save('result.xlsx')
